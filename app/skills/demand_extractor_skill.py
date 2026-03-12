"""Demand extractor skill - extract requirements from prototype pages"""

import re
import os
from typing import Dict, Any, List, Optional
from .base import BaseSkill, SkillResult
from ..core.session import SessionState
from playwright.sync_api import sync_playwright


class DemandExtractorSkill(BaseSkill):
    """需求提取器 - 从原型页面提取需求内容"""
    
    name = "extract_demand"
    description = "从原型页面提取需求内容并生成测试用例"
    triggers = ["提取需求", "解析需求", "提取页面"]
    
    parameters = {
        "selection": {"required": False, "description": "选择的页面名称"},
    }
    
    def execute(self, params: Dict[str, Any], session: SessionState) -> SkillResult:
        flat_pages = session.analyzed_pages
        
        if not flat_pages:
            return SkillResult(success=False, error="没有可用的页面数据", suggestion="请先分析原型链接")
        
        selection = params.get("selection", "")
        
        if not selection:
            return SkillResult(
                success=False,
                error="请选择要提取需求的页面",
                suggestion=f"可用页面：{', '.join([p['name'] for p in flat_pages[:5]])}..."
            )
        
        # 筛选页面
        keywords = re.split(r'[,，、\s]+', selection)
        keywords = [k.strip() for k in keywords if k.strip()]
        
        selected_pages = []
        for p in flat_pages:
            for keyword in keywords:
                if keyword.lower() in p['name'].lower():
                    if p not in selected_pages:
                        selected_pages.append(p)
                    break
        
        if not selected_pages:
            return SkillResult(success=False, error=f"没有找到匹配「{selection}」的页面")
        
        url = session.current_url
        if not url:
            return SkillResult(success=False, error="URL不存在")
        
        # 提取需求内容
        demands = self._extract_demands_playwright(url, selected_pages)
        
        if not demands:
            # 如果提取失败，使用简化方式
            demands = self._simple_extract(selected_pages)
        
        if not demands:
            return SkillResult(success=False, error="未能提取到需求内容")
        
        # 使用LLM梳理需求
        analyzed_demands = self._analyze_demands(demands)
        
        # 生成测试用例
        test_cases = self._generate_test_cases(analyzed_demands)
        
        # 导出到临时目录
        export_path = self._export_test_cases(test_cases)
        
        # 更新会话
        session.test_cases = test_cases
        
        return SkillResult(
            success=True,
            data={
                "demands": analyzed_demands,
                "test_cases": test_cases,
                "total": len(test_cases),
                "export_path": export_path,
                "selected_pages": [p['name'] for p in selected_pages]
            },
            message=f"已提取 {len(demands)} 个页面需求，生成 {len(test_cases)} 条测试用例\n导出文件：{export_path}"
        )
    
    def _extract_demands_playwright(self, url: str, pages: List[Dict]) -> List[Dict]:
        """使用Playwright提取页面需求内容"""
        demands = []
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 1920, "height": 1080})
            
            try:
                page.goto(url, timeout=60000, wait_until="networkidle")
                page.wait_for_timeout(8000)  # 等待iframe加载
                
                # 尝试切换到iframe
                try:
                    main_frame = page.frame_locator('#mainFrame, iframe[name="mainFrame"]').first
                    if main_frame:
                        page = main_frame
                except:
                    pass
                
                for page_info in pages:
                    page_name = page_info.get('name', '')
                    
                    try:
                        # 尝试点击sitemap中的页面
                        clicked = self._try_click_sitemap_item(page, page_name)
                        
                        if clicked:
                            page.wait_for_timeout(2000)
                            self._scroll_page(page)
                            content = self._get_page_content(page)
                            
                            if content.get('text') or content.get('elements'):
                                demands.append({
                                    "page_name": page_name,
                                    "content": content
                                })
                    except Exception as e:
                        print(f"Error extracting {page_name}: {e}")
                
            finally:
                browser.close()
        
        return demands
    
    def _try_click_sitemap_item(self, page, page_name: str) -> bool:
        """尝试点击sitemap中的页面项"""
        selectors = [
            f'text="{page_name}"',
            f'text={page_name}',
            f'[class*="sitemap"] >> text="{page_name}"',
            f'a:has-text("{page_name}")',
            f'span:has-text("{page_name}")',
        ]
        
        for selector in selectors:
            try:
                elem = page.locator(selector).first
                if elem.count() > 0:
                    elem.click(timeout=3000)
                    return True
            except:
                continue
        
        return False
    
    def _scroll_page(self, page):
        """滚动页面加载所有内容"""
        try:
            for i in range(5):
                page.evaluate(f"window.scrollBy(0, 500)")
                page.wait_for_timeout(500)
            page.evaluate("window.scrollTo(0, 0)")
        except:
            pass
    
    def _get_page_content(self, page) -> Dict:
        """获取页面内容"""
        try:
            text = page.evaluate("() => document.body?.innerText || ''")
            
            elements = page.evaluate("""() => {
                const items = [];
                
                // 按钮
                document.querySelectorAll('button, [role="button"], .btn, [class*="button"]').forEach(el => {
                    const text = el.innerText?.trim();
                    if (text) items.push({type: 'button', text});
                });
                
                // 输入框
                document.querySelectorAll('input, textarea, select').forEach(el => {
                    const label = el.placeholder || el.name || el.id;
                    if (label) items.push({type: 'input', label, inputType: el.type});
                });
                
                // 标题
                document.querySelectorAll('h1, h2, h3, h4, h5, h6').forEach(el => {
                    const text = el.innerText?.trim();
                    if (text) items.push({type: 'heading', text, level: el.tagName});
                });
                
                return items.slice(0, 30);
            }""")
            
            return {"text": text, "elements": elements}
        except Exception as e:
            return {"text": "", "elements": [], "error": str(e)}
    
    def _simple_extract(self, pages: List[Dict]) -> List[Dict]:
        """简化提取（当Playwright提取失败时）"""
        demands = []
        
        for p in pages:
            page_name = p.get('name', '')
            
            # 基于页面名称推断需求
            demands.append({
                "page_name": page_name,
                "content": {
                    "text": f"页面名称：{page_name}",
                    "elements": []
                }
            })
        
        return demands
    
    def _analyze_demands(self, demands: List[Dict]) -> List[Dict]:
        """使用LLM分析需求"""
        if not self.llm:
            return demands
        
        analyzed = []
        
        for demand in demands:
            content = demand.get("content", {})
            text = content.get("text", "")
            elements = content.get("elements", [])
            page_name = demand.get("page_name", "")
            
            prompt = f"""分析以下原型页面，提取需求信息：

页面名称：{page_name}

页面内容：
{text[:1500]}

页面元素：
{self._format_elements(elements[:15])}

请输出JSON格式的需求分析：
{{
  "function": "功能描述",
  "inputs": ["输入项列表"],
  "outputs": ["输出项列表"],
  "rules": ["业务规则列表"],
  "exceptions": ["异常处理列表"]
}}
"""
            
            try:
                from ..llm import Message, MessageRole
                response = self.llm.chat([Message(role=MessageRole.USER, content=prompt)])
                
                json_match = re.search(r'\{.*\}', response, re.DOTALL)
                if json_match:
                    import json
                    analysis = json.loads(json_match.group())
                else:
                    analysis = {"raw": response}
                
                demand["analysis"] = analysis
            except Exception as e:
                demand["analysis"] = {"error": str(e)}
            
            analyzed.append(demand)
        
        return analyzed
    
    def _format_elements(self, elements: List[Dict]) -> str:
        """格式化元素列表"""
        lines = []
        for el in elements:
            el_type = el.get("type", "")
            if el_type == "button":
                lines.append(f"- 按钮: {el.get('text', '')}")
            elif el_type == "input":
                lines.append(f"- 输入框: {el.get('label', '')}")
            elif el_type == "heading":
                lines.append(f"- 标题: {el.get('text', '')}")
        
        return "\n".join(lines) if lines else "无"
    
    def _generate_test_cases(self, demands: List[Dict]) -> List[Dict]:
        """生成测试用例"""
        test_cases = []
        tc_id = 1
        
        for demand in demands:
            page_name = demand.get("page_name", "")
            analysis = demand.get("analysis", {})
            
            # 正向用例
            test_cases.append({
                "id": f"TC_{tc_id:03d}",
                "title": f"正向-{page_name}功能验证",
                "preconditions": "用户已登录",
                "steps": f"1. 进入{page_name}页面\n2. 按正常流程操作\n3. 验证结果",
                "expected_results": "功能正常执行",
                "priority": "P1",
                "type": "positive",
                "page": page_name
            })
            tc_id += 1
            
            # 逆向用例
            test_cases.append({
                "id": f"TC_{tc_id:03d}",
                "title": f"逆向-{page_name}异常输入",
                "preconditions": "用户已登录",
                "steps": f"1. 进入{page_name}页面\n2. 输入非法数据\n3. 提交操作",
                "expected_results": "系统提示错误",
                "priority": "P2",
                "type": "negative",
                "page": page_name
            })
            tc_id += 1
            
            # 边界用例
            test_cases.append({
                "id": f"TC_{tc_id:03d}",
                "title": f"边界-{page_name}边界值测试",
                "preconditions": "用户已登录",
                "steps": f"1. 进入{page_name}页面\n2. 输入边界值\n3. 验证系统处理",
                "expected_results": "系统正确处理边界情况",
                "priority": "P2",
                "type": "boundary",
                "page": page_name
            })
            tc_id += 1
        
        return test_cases
    
    def _export_test_cases(self, test_cases: List[Dict]) -> str:
        """导出测试用例到/tmp目录"""
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        tmp_dir = "/tmp/demandtest"
        os.makedirs(tmp_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_cases_{timestamp}.xlsx"
        filepath = os.path.join(tmp_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        headers = ["ID", "标题", "前置条件", "操作步骤", "预期结果", "优先级", "类型", "页面"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        
        for row, tc in enumerate(test_cases, 2):
            values = [
                tc.get("id", ""),
                tc.get("title", ""),
                tc.get("preconditions", ""),
                tc.get("steps", ""),
                tc.get("expected_results", ""),
                tc.get("priority", ""),
                tc.get("type", ""),
                tc.get("page", "")
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        column_widths = [15, 35, 25, 40, 30, 10, 12, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        ws.freeze_panes = "A2"
        wb.save(filepath)
        
        return filepath
