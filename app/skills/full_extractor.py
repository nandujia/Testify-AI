#!/usr/bin/env python3
"""
完整的需求提取流程：
1. 使用modao_crawler获取页面列表和screen ID
2. 根据页面名称导航到目标页面
3. 截图+OCR提取内容
4. 需求分析
5. 生成测试用例
"""

import os
import sys
import json
import base64
import re
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass
from datetime import datetime

import requests
from playwright.sync_api import sync_playwright

# 添加crawler路径
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "demand-test-platform/backend/app/services/crawler"))
from modao_crawler import ModaoCrawler


@dataclass
class PageContent:
    """页面内容"""
    name: str
    screen_id: str
    text: str
    screenshots: List[str]
    analysis: str = ""


class FullDemandExtractor:
    """完整需求提取器"""
    
    def __init__(self, output_dir: str = "/tmp/demandtest"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.screenshots_dir = self.output_dir / "screenshots"
        self.screenshots_dir.mkdir(exist_ok=True)
    
    def extract(
        self, 
        url: str, 
        page_names: List[str] = None,
        use_ocr: bool = True
    ) -> Dict:
        """提取需求主流程"""
        
        print(f"\n{'='*70}")
        print(f"需求提取流程")
        print(f"{'='*70}")
        print(f"URL: {url}")
        print(f"目标页面: {page_names or '全部'}")
        print(f"{'='*70}\n")
        
        results = {
            "url": url,
            "page_names": page_names,
            "pages": [],
            "contents": [],
            "test_cases": []
        }
        
        # Step 1: 获取页面列表
        print("Step 1: 获取页面列表...")
        crawler = ModaoCrawler()
        crawl_result = crawler.crawl(url)
        
        if not crawl_result.get("success"):
            print(f"  错误: {crawl_result.get('error', '未知错误')}")
            return results
        
        all_pages = self._flatten_pages(crawl_result.get("pages", []))
        print(f"  找到 {len(all_pages)} 个页面")
        
        results["pages"] = all_pages
        
        # Step 2: 筛选目标页面
        if page_names:
            target_pages = [
                p for p in all_pages 
                if any(name.lower() in p.get("name", "").lower() for name in page_names)
            ]
        else:
            target_pages = all_pages[:5]  # 默认只提取前5个
        
        print(f"  目标页面: {len(target_pages)} 个")
        
        for p in target_pages:
            print(f"    - {p.get('name')} (screen: {p.get('screen_id')})")
        
        # Step 3: 提取每个页面的内容
        print(f"\nStep 2: 提取页面内容...")
        
        for i, page_info in enumerate(target_pages):
            print(f"\n  [{i+1}/{len(target_pages)}] {page_info.get('name')}")
            
            content = self._extract_page_content(url, page_info, use_ocr)
            results["contents"].append(content)
            
            # Step 4: 需求分析
            if content.text:
                print(f"    分析需求...")
                content.analysis = self._analyze_demand(content.name, content.text)
                
                # Step 5: 生成测试用例
                print(f"    生成测试用例...")
                test_cases = self._generate_test_cases(content)
                results["test_cases"].extend(test_cases)
        
        # Step 6: 导出结果
        print(f"\nStep 3: 导出结果...")
        export_file = self._export_results(results)
        results["export_file"] = str(export_file)
        
        print(f"\n{'='*70}")
        print(f"提取完成！")
        print(f"  页面数: {len(all_pages)}")
        print(f"  提取数: {len(target_pages)}")
        print(f"  用例数: {len(results['test_cases'])}")
        print(f"  导出文件: {export_file}")
        print(f"{'='*70}\n")
        
        return results
    
    def _flatten_pages(self, pages: List[Dict]) -> List[Dict]:
        """展平页面列表"""
        result = []
        
        def _flatten(items: List[Dict], parent_name: str = ""):
            for item in items:
                name = item.get("name", "")
                screen_id = item.get("screen_id", "")
                
                result.append({
                    "name": name,
                    "screen_id": screen_id,
                    "parent": parent_name
                })
                
                children = item.get("children", [])
                if children:
                    _flatten(children, name)
        
        _flatten(pages)
        return result
    
    def _extract_page_content(
        self, 
        base_url: str, 
        page_info: Dict,
        use_ocr: bool
    ) -> PageContent:
        """提取单个页面内容"""
        
        screen_id = page_info.get("screen_id", "")
        page_name = page_info.get("name", "Unknown")
        
        # 构造页面URL
        if '?' in base_url:
            page_url = f"{base_url}&screen={screen_id}"
        else:
            page_url = f"{base_url}?s=0&screen={screen_id}"
        
        screenshots = []
        text = ""
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 1920, "height": 1080})
            
            # 访问页面
            page.goto(page_url, timeout=60000)
            page.wait_for_load_state('networkidle')
            page.wait_for_timeout(5000)
            
            # 获取iframe
            iframe_url = page.evaluate("""() => {
                const iframe = document.querySelector('iframe');
                return iframe ? iframe.src : null;
            }""")
            
            if iframe_url:
                # 打开iframe
                iframe_page = browser.new_page(viewport={"width": 1920, "height": 1080})
                iframe_page.goto(iframe_url, timeout=30000)
                iframe_page.wait_for_load_state('networkidle')
                iframe_page.wait_for_timeout(3000)
                
                # 获取文本内容
                text = iframe_page.evaluate("document.body.innerText")
                
                # 滚动截图
                scroll_height = iframe_page.evaluate("document.body.scrollHeight")
                viewport_height = iframe_page.evaluate("window.innerHeight")
                scroll_step = 800
                max_scrolls = min(15, (scroll_height // scroll_step) + 2)
                
                print(f"    文本长度: {len(text)}, 页面高度: {scroll_height}")
                
                for i in range(max_scrolls):
                    screenshot_file = self.screenshots_dir / f"{page_name}_{i}.png"
                    iframe_page.screenshot(path=str(screenshot_file))
                    screenshots.append(str(screenshot_file))
                    
                    iframe_page.evaluate(f"window.scrollBy(0, {scroll_step})")
                    iframe_page.wait_for_timeout(500)
                    
                    current_scroll = iframe_page.evaluate("window.scrollY")
                    if current_scroll + viewport_height >= scroll_height - 100:
                        break
                
                iframe_page.close()
            
            browser.close()
        
        # OCR处理（如果需要且文本较少）
        if use_ocr and len(text) < 100 and screenshots:
            print(f"    文本较少，尝试OCR...")
            text = self._ocr_screenshots(screenshots)
        
        return PageContent(
            name=page_name,
            screen_id=screen_id,
            text=text,
            screenshots=screenshots
        )
    
    def _ocr_screenshots(self, screenshots: List[str]) -> str:
        """OCR识别截图"""
        texts = []
        
        for i, screenshot_path in enumerate(screenshots[:5]):  # 最多OCR前5张
            print(f"      OCR [{i+1}/{len(screenshots[:5])}]")
            
            with open(screenshot_path, "rb") as f:
                image_base64 = base64.b64encode(f.read()).decode()
            
            try:
                text = self._call_vision_api(image_base64)
                if text:
                    texts.append(text)
            except Exception as e:
                print(f"        OCR失败: {e}")
        
        return "\n\n".join(texts)
    
    def _call_vision_api(self, image_base64: str) -> str:
        """调用视觉API"""
        # 尝试GLM-4V
        url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
        api_key = os.getenv("ZHIPU_API_KEY", "")
        
        if not api_key:
            # 返回模拟结果
            return ""
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": "glm-4v-flash",
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "识别图片中的所有文字，包括标题、正文、按钮文字等。按从上到下、从左到右顺序输出。"
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{image_base64}"}
                        }
                    ]
                }
            ],
            "max_tokens": 2000
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        result = response.json()
        
        return result.get("choices", [{}])[0].get("message", {}).get("content", "")
    
    def _analyze_demand(self, page_name: str, text: str) -> str:
        """分析需求"""
        prompt = f"""分析以下需求内容，提取关键信息。

页面名称: {page_name}

需求内容:
{text[:3000]}

请按以下格式输出：

## 功能模块
列出主要功能模块

## 输入项
列出输入字段

## 输出项
列出输出内容

## 业务规则
列出关键规则

## 异常场景
列出可能的异常
"""
        
        # 调用LLM
        url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
        api_key = os.getenv("ZHIPU_API_KEY", "")
        
        if not api_key:
            return "需要配置API Key才能分析需求"
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": "glm-4-flash",
            "messages": [
                {"role": "system", "content": "你是专业的需求分析师。"},
                {"role": "user", "content": prompt}
            ]
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        result = response.json()
        
        return result.get("choices", [{}])[0].get("message", {}).get("content", "")
    
    def _generate_test_cases(self, content: PageContent) -> List[Dict]:
        """生成测试用例"""
        prompt = f"""根据以下需求分析，生成测试用例。

页面: {content.name}

需求分析:
{content.analysis[:2000]}

生成测试用例（JSON数组格式）：
[
  {{
    "id": "TC_001",
    "title": "正向-XXX功能验证",
    "preconditions": "前置条件",
    "steps": "操作步骤",
    "expected_results": "预期结果",
    "priority": "P0",
    "type": "positive"
  }}
]

覆盖类型：positive（正向）、negative（逆向）、boundary（边界）、exception（异常）
"""
        
        url = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
        api_key = os.getenv("ZHIPU_API_KEY", "")
        
        if not api_key:
            return []
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": "glm-4-flash",
            "messages": [
                {"role": "system", "content": "你是专业的测试工程师。"},
                {"role": "user", "content": prompt}
            ]
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=60)
            result = response.json()
            
            content_str = result.get("choices", [{}])[0].get("message", {}).get("content", "[]")
            
            # 提取JSON
            json_match = re.search(r'\[[\s\S]*\]', content_str)
            if json_match:
                test_cases = json.loads(json_match.group())
                for tc in test_cases:
                    tc["page"] = content.name
                return test_cases
        except Exception as e:
            print(f"        生成用例失败: {e}")
        
        return []
    
    def _export_results(self, results: Dict) -> Path:
        """导出结果"""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        
        # Sheet 1: 页面列表
        ws_pages = wb.active
        ws_pages.title = "页面列表"
        ws_pages.append(["页面名称", "Screen ID", "父页面"])
        for page in results["pages"]:
            ws_pages.append([page.get("name", ""), page.get("screen_id", ""), page.get("parent", "")])
        
        # Sheet 2: 需求内容
        ws_content = wb.create_sheet("需求内容")
        ws_content.append(["页面名称", "文本内容", "需求分析"])
        for content in results["contents"]:
            ws_content.append([content.name, content.text[:2000], content.analysis[:1000]])
        
        # Sheet 3: 测试用例
        ws_cases = wb.create_sheet("测试用例")
        headers = ["用例ID", "用例标题", "前置条件", "操作步骤", "预期结果", "优先级", "用例类型", "页面"]
        ws_cases.append(headers)
        
        for tc in results["test_cases"]:
            ws_cases.append([
                tc.get("id", ""),
                tc.get("title", ""),
                tc.get("preconditions", ""),
                tc.get("steps", ""),
                tc.get("expected_results", ""),
                tc.get("priority", ""),
                tc.get("type", ""),
                tc.get("page", "")
            ])
        
        # 保存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.output_dir / f"需求提取_{timestamp}.xlsx"
        wb.save(filepath)
        
        return filepath


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="完整需求提取器")
    parser.add_argument("url", help="墨刀页面URL")
    parser.add_argument("pages", nargs="*", help="页面名称")
    parser.add_argument("--output", "-o", default="/tmp/demandtest", help="输出目录")
    
    args = parser.parse_args()
    
    extractor = FullDemandExtractor(output_dir=args.output)
    results = extractor.extract(args.url, args.pages if args.pages else None)
    
    if results.get("export_file"):
        print(f"\n导出文件: {results['export_file']}")


if __name__ == "__main__":
    main()
