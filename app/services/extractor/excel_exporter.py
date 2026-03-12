"""
Excel 导出服务

将测试用例导出为标准 Excel 格式，兼容 TestRail 等测试管理系统
"""

import os
from typing import List, Dict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("请安装 openpyxl: pip install openpyxl")


class ExcelExporter:
    """Excel 导出器"""
    
    # 列定义
    COLUMNS = [
        ("用例编号", 15),
        ("用例标题", 35),
        ("前置条件", 25),
        ("测试步骤", 45),
        ("预期结果", 45),
        ("优先级", 10),
        ("测试类型", 12),
        ("备注", 15),
    ]
    
    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export(self, test_cases: List[Dict], filename: str = None) -> str:
        """
        导出测试用例到 Excel
        
        Args:
            test_cases: 测试用例列表
            filename: 文件名（可选）
            
        Returns:
            导出文件路径
        """
        if not test_cases:
            raise ValueError("测试用例列表为空")
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"测试用例_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 写入表头
        self._write_header(ws)
        
        # 写入数据
        self._write_data(ws, test_cases)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def _write_header(self, ws):
        """写入表头"""
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, (col_name, col_width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _write_data(self, ws, test_cases: List[Dict]):
        """写入数据"""
        # 样式定义
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 字段映射
        field_mapping = {
            "用例编号": "id",
            "用例标题": "title",
            "前置条件": "preconditions",
            "测试步骤": "steps",
            "预期结果": "expected_results",
            "优先级": "priority",
            "测试类型": "type",
            "备注": "remarks",
        }
        
        for row_idx, test_case in enumerate(test_cases, 2):
            for col_idx, (col_name, _) in enumerate(self.COLUMNS, 1):
                field = field_mapping.get(col_name)
                value = test_case.get(field, "") if field else ""
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
            
            # 设置行高
            ws.row_dimensions[row_idx].height = 60


# ==================== 统一入口 ====================

def export_to_excel(test_cases: List[Dict], filename: str = None, output_dir: str = "./exports") -> str:
    """
    导出测试用例到 Excel
    
    Args:
        test_cases: 测试用例列表
        filename: 文件名
        output_dir: 输出目录
        
    Returns:
        导出文件路径
    """
    exporter = ExcelExporter(output_dir)
    return exporter.export(test_cases, filename)
